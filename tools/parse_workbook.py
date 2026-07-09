import openpyxl, re, json, sys
from openpyxl.utils import get_column_letter as gcl

SRC = sys.argv[1] if len(sys.argv)>1 else "/root/.claude/uploads/c419d4aa-8671-5236-8e6c-7c2f4782364f/c556e32e-Blood_Bank_Daily_Statistics_JUNE.xlsx"
wb = openpyxl.load_workbook(SRC, data_only=True)

MONTHS=[('JAN',1),('FEB',2),('MARCH',3),('MAR',3),('APRIL',4),('APR',5-1),('MAY',5),('JUNE',6),('JUN',6),
        ('JULY',7),('JUL',7),('AGU',8),('AUG',8),('SEPT',9),('SEPTAMBER',9),('SEP',9),('OCTO',10),('OCT',10),
        ('NOVEMBR',11),('NOV',12-1),('DEC',12)]
# clean explicit map to avoid the APR/NOV typos above
MONTHMAP=[('SEPTAMBER',9),('NOVEMBR',11),('OCTO',10),('MARCH',3),('APRIL',4),('AUGUST',8),('AGU',8),
          ('JANUARY',1),('FEBRUARY',2),('JAN',1),('FEB',2),('MAR',3),('APR',4),('MAY',5),
          ('JUNE',6),('JULY',7),('JUN',6),('JUL',7),('AUG',8),('SEP',9),('OCT',10),('NOV',11),('DEC',12),('DECEMBER',12)]

def detect(title):
    u=title.upper()
    mo=None
    for k,v in MONTHMAP:
        if k in u: mo=v; break
    m=re.search(r'(20\d\d)',u)
    yr=int(m.group(1)) if m else 2024   # 'Blood components issued' family = 2024
    return mo,yr

def norm(s): return re.sub(r'\s+',' ',str(s)).strip() if s is not None else ''
def nu(s): return norm(s).upper()

# Canonical section keys
def section_key(label):
    u=nu(label)
    if u.startswith('ISSUING PRBC'): return ('issue','PRBC')
    if u.startswith('ISSUING FFP'): return ('issue','FFP')
    if u.startswith('ISSUING CRYO'): return ('issue','CRYO')
    if u.startswith('ISSUING PLATELET'): return ('issue','Platelets')
    if u.startswith('RECEIVED CROSS'): return ('received',None)
    if u.startswith('RETURNED FROM WARD'): return ('returned_ward',None)
    if 'RETURNED FROM ARH' in u: return ('returned_ash',None)
    if 'DAILY INVENTORY' in u: return ('inventory',None)
    return None

WARD_CANON={
 'REHAB ICU 4TH FLOOR':'Rehab ICU 4th floor','LTC 3RD FLOOR':'LTC 3rd floor','LTC 5TH FLOOR':'LTC 5th floor',
 'LTC 6TH FLOOR':'LTC 6th floor','LTC 7TH FLOOR':'LTC 7th floor','L8TH-MEDICAL FLOOR':'L8th-medical floor',
 'LTC 9TH FLOOR':'LTC 9th floor','12 FLOOR':'12 floor','13 FLOOR':'13 floor','REHAB 14 FLOOR':'Rehab 14 floor',
 'HDU':'HDU','DHDU':'HDU'}
def ward_canon(label):
    u=nu(label)
    return WARD_CANON.get(u, norm(label))
COMP_CANON={'PRBC':'PRBC','PLATELETS':'Platelets','FFP':'FFP','CRYO.':'CRYO','CRYO':'CRYO'}
def comp_canon(label):
    return COMP_CANON.get(nu(label), norm(label))

months=[]
for ws in wb.worksheets:
    mo,yr=detect(ws.title)
    if not mo:
        print("SKIP (no month):", ws.title); continue
    # find day-column map from any section header row: columns with int 1..31, and Total
    entries=[]  # list of dict
    r=1
    maxr=ws.max_row; maxc=ws.max_column
    cur=None
    daily_family = yr>=2025  # current family has per-day; 2024 issued family is totals only
    while r<=maxr:
        b=ws.cell(r,2).value
        sk=section_key(b) if b is not None and nu(b) else None
        if sk:
            # build day map from this header row
            daymap={}; total_col=None
            for c in range(3,maxc+1):
                hv=ws.cell(r,c).value
                if isinstance(hv,(int,float)) and 1<=hv<=31:
                    daymap[int(hv)]=c
                elif isinstance(hv,str) and hv.strip().upper()=='TOTAL':
                    total_col=c
            if total_col is None:
                # fallback: AH col 34
                total_col=34
            cur={'sk':sk,'daymap':daymap,'total_col':total_col,'header_row':r}
            r+=1
            continue
        if cur and b is not None and nu(b):
            # data row: labelled ward/component
            skind,scomp=cur['sk']
            label=norm(b)
            if skind in ('returned_ward','returned_ash','inventory'):
                rowlabel=comp_canon(b)   # rows are components
            else:
                rowlabel=ward_canon(b)   # rows are wards
            days={}
            for d,c in cur['daymap'].items():
                v=ws.cell(r,c).value
                if isinstance(v,(int,float)) and v!=0:
                    days[d]=v
            tv=ws.cell(r,cur['total_col']).value
            total=tv if isinstance(tv,(int,float)) else None
            entries.append({'section':skind,'component':scomp,'row':rowlabel,'days':days,'total':total})
        r+=1
    months.append({'title':ws.title,'year':yr,'month':mo,'daily':daily_family,'entries':entries})

# dedupe by (year,month): keep first, but warn on collisions
seen={}
for m in months:
    k=(m['year'],m['month'])
    seen.setdefault(k,[]).append(m['title'])
for k,v in sorted(seen.items()):
    if len(v)>1: print("COLLISION",k,v)

out={'wards':['Rehab ICU 4th floor','LTC 3rd floor','LTC 5th floor','LTC 6th floor','LTC 7th floor',
      'L8th-medical floor','LTC 9th floor','12 floor','13 floor','Rehab 14 floor','HDU'],
     'components':['PRBC','FFP','CRYO','Platelets'],
     'months':months}
json.dump(out, open("/home/user/BBData/tools/seed.json","w"), indent=1)
print("months:",len(months))
print("total entries:", sum(len(m['entries']) for m in months))
# sanity print one
import statistics
for m in months[:2]:
    print(m['title'],m['year'],m['month'],'daily' if m['daily'] else 'totals','entries',len(m['entries']))
