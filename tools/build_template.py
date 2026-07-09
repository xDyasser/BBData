#!/usr/bin/env python3
"""Build a professionally styled Excel template (with native charts + formulas)
for the Blood Bank app, pre-populated from tools/seed.json, then base64-embed it
as template-embed.js. The web app injects live values into this template's cells
via JSZip XML surgery at export time, preserving all styling and charts.

Layout is fixed so the app can target cells deterministically:
  Summary  : period B2, months B3, generated B4, activity B7:B11,
             components A14:A17 (labels), B14:B17 (units), C14:C17 (share formula)
  By Year  : years A3:A(2+Y), data B3:E(2+Y); reserved 12 rows (3..14); Total row 15
  By Ward  : wards A3:A13 (11 fixed), data B3:E13, F total, G share; Total row 14
  By Month : months A3:A(2+M), data B3:E, F received, G returned, H total;
             reserved 120 rows (3..122)
"""
import json, os, base64
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_js = open(os.path.join(ROOT, "data-seed.js")).read()
seed = json.loads(_js[_js.index("{"):_js.rindex("}")+1])
COMPS = seed["components"]            # PRBC, FFP, CRYO, Platelets
WARDS = seed["wards"]                 # 11 wards
DATA  = seed["data"]
KEYS  = sorted(DATA.keys())
YEARS = sorted({k[:4] for k in KEYS})

MONTH_ABBR = ["","Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
def klabel(k): y,m=k.split("-"); return f"{MONTH_ABBR[int(m)]} {y}"

YEARS_RESERVED = 12
MONTHS_RESERVED = 120

# ---- aggregation from seed ----
def issue_month(k, comp=None, ward=None):
    M = DATA[k]["issue"]; t=0
    comps = [comp] if comp else COMPS
    for c in comps:
        cw = M.get(c, {})
        wards = [ward] if ward else cw.keys()
        for w in wards:
            for d,v in cw.get(w, {}).items(): t += v
    return t
def section_month(k, sec):
    if sec=="issue": return issue_month(k)
    blk = DATA[k][sec]; t=0
    for r in blk.values():
        for v in r.values(): t+=v
    return t
def year_comp(y, c): return sum(issue_month(k, c) for k in KEYS if k.startswith(y))
def ward_comp(w, c): return sum(issue_month(k, c, w) for k in KEYS)

# ---- colors / styles ----
RED="C0392B"; INK="1F2933"; LINE="D9D6CE"; BAND="F5F3EF"; TOTALF="EFEBE3"
COMPFILL={"PRBC":"E34948","FFP":"EDA100","CRYO":"2A78D6","Platelets":"1BAF7A"}
thin = Side(style="thin", color=LINE)
BORDER = Border(left=thin,right=thin,top=thin,bottom=thin)
NUMFMT = '#,##0'
PCTFMT = '0.0"%"'

wb = openpyxl.Workbook()

def style_header_cell(c, fill=RED, color="FFFFFF"):
    c.font=Font(bold=True, color=color, size=11)
    c.fill=PatternFill("solid", fgColor=fill)
    c.alignment=Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border=BORDER
def title(ws, text, span):
    ws.merge_cells(f"A1:{get_column_letter(span)}1")
    c=ws["A1"]; c.value=text
    c.font=Font(bold=True, color="FFFFFF", size=14)
    c.fill=PatternFill("solid", fgColor=RED)
    c.alignment=Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height=26
def data_cell(c, num=True, total=False, band=False):
    c.border=BORDER
    c.alignment=Alignment(horizontal="right" if num else "left", vertical="center", indent=0 if num else 1)
    if num: c.number_format=NUMFMT
    if total: c.font=Font(bold=True); c.fill=PatternFill("solid", fgColor=TOTALF)
    elif band: c.fill=PatternFill("solid", fgColor=BAND)

def finish_chart(ch, xt=None, yt=None):
    # keep the title above the plot (not overlaying the bars) and show axis titles
    if ch.title is not None:
        try: ch.title.overlay = False
        except Exception: pass
    if hasattr(ch, "x_axis"):
        ch.x_axis.delete = False
        ch.y_axis.delete = False
        if xt is not None: ch.x_axis.title = xt
        if yt is not None: ch.y_axis.title = yt
        try:
            ch.x_axis.title.overlay = False
            ch.y_axis.title.overlay = False
        except Exception: pass
    ch.legend.position = "b"

# ================= Summary =================
ws = wb.active; ws.title="Summary"
title(ws, "Blood Bank Statistics — Summary", 4)
labels=[("A2","Reporting period"),("A3","Months on record"),("A4","Generated")]
for ref,t in labels:
    ws[ref]=t; ws[ref].font=Font(bold=True, color=INK)
ws["B2"]=f"{klabel(KEYS[0])} – {klabel(KEYS[-1])}"
ws["B3"]=len(KEYS)
ws["B4"]="(set on export)"
# activity table
ws["A6"]="Activity"; ws["B6"]="Total units"
style_header_cell(ws["A6"]); style_header_cell(ws["B6"])
acts=[("Units issued",section_month_all:=sum(issue_month(k) for k in KEYS)),
      ("Cross-matched received",sum(section_month(k,"received") for k in KEYS)),
      ("Returned from ward",sum(section_month(k,"returned_ward") for k in KEYS)),
      ("Returned ARH→ASH",sum(section_month(k,"returned_ash") for k in KEYS)),
      ("Inventory from ASH",sum(section_month(k,"inventory") for k in KEYS))]
for i,(lab,val) in enumerate(acts):
    r=7+i
    ws[f"A{r}"]=lab; data_cell(ws[f"A{r}"], num=False, band=(i%2==1))
    ws[f"B{r}"]=val; data_cell(ws[f"B{r}"], band=(i%2==1))
# component share table
ws["A13"]="Component"; ws["B13"]="Units issued"; ws["C13"]="Share"
for col in ("A13","B13","C13"): style_header_cell(ws[col])
grand=sum(sum(year_comp(y,c) for y in YEARS) for c in COMPS) or 1
for i,c in enumerate(COMPS):
    r=14+i
    ws[f"A{r}"]=c
    ws[f"A{r}"].font=Font(bold=True, color="FFFFFF")
    ws[f"A{r}"].fill=PatternFill("solid", fgColor=COMPFILL[c])
    ws[f"A{r}"].border=BORDER; ws[f"A{r}"].alignment=Alignment(indent=1, vertical="center")
    ws[f"B{r}"]=sum(year_comp(y,c) for y in YEARS); data_cell(ws[f"B{r}"])
    ws[f"C{r}"]=f"=B{r}/SUM($B$14:$B$17)*100"; data_cell(ws[f"C{r}"]); ws[f"C{r}"].number_format=PCTFMT
ws.column_dimensions["A"].width=24; ws.column_dimensions["B"].width=16; ws.column_dimensions["C"].width=10
pie=PieChart(); pie.title="Component share"; pie.height=7; pie.width=11
pie.add_data(Reference(ws,min_col=2,min_row=14,max_row=17)); pie.set_categories(Reference(ws,min_col=1,min_row=14,max_row=17))
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
ser=pie.series[0]
for i,c in enumerate(COMPS):
    dp=DataPoint(idx=i); dp.graphicalProperties.solidFill=COMPFILL[c]; ser.data_points.append(dp)
finish_chart(pie)
ws.add_chart(pie,"E6")
ws.freeze_panes="A2"

# ================= By Year =================
wy = wb.create_sheet("By Year")
title(wy, "Units issued by year & component", 6)
hdr=["Year"]+COMPS+["Total"]
for j,h in enumerate(hdr):
    c=wy.cell(2,1+j,h)
    style_header_cell(c, fill=(COMPFILL[h] if h in COMPFILL else RED))
for i in range(YEARS_RESERVED):
    r=3+i
    y = YEARS[i] if i < len(YEARS) else None
    wy.cell(r,1, y if y else None); data_cell(wy.cell(r,1), num=False, band=(i%2==1))
    for j,c in enumerate(COMPS):
        cell=wy.cell(r,2+j, (year_comp(y,c) if y else None)); data_cell(cell, band=(i%2==1))
    tot=wy.cell(r,6, f"=SUM(B{r}:E{r})" if y else None); data_cell(tot, total=True)
tr=3+YEARS_RESERVED
wy.cell(tr,1,"Total"); data_cell(wy.cell(tr,1), num=False, total=True)
for j in range(len(COMPS)+1):
    col=get_column_letter(2+j)
    wy.cell(tr,2+j, f"=SUM({col}3:{col}{tr-1})"); data_cell(wy.cell(tr,2+j), total=True)
wy.column_dimensions["A"].width=10
for j in range(len(COMPS)+1): wy.column_dimensions[get_column_letter(2+j)].width=12
wy.freeze_panes="B3"
bar=BarChart(); bar.type="col"; bar.grouping="clustered"; bar.title="Yearly issuance by component"
bar.height=9; bar.width=18
bar.add_data(Reference(wy,min_col=2,max_col=5,min_row=2,max_row=2+len(YEARS)), titles_from_data=True)
bar.set_categories(Reference(wy,min_col=1,min_row=3,max_row=2+len(YEARS)))
for i,c in enumerate(COMPS): bar.series[i].graphicalProperties.solidFill=COMPFILL[c]
finish_chart(bar, xt="Year", yt="Units issued")
wy.add_chart(bar,"H2")

# ================= By Ward =================
ww = wb.create_sheet("By Ward")
title(ww, "Units issued by ward & component", 7)
hdr=["Ward / Floor"]+COMPS+["Total","Share"]
for j,h in enumerate(hdr):
    c=ww.cell(2,1+j,h); style_header_cell(c, fill=(COMPFILL[h] if h in COMPFILL else RED))
for i,w in enumerate(WARDS):
    r=3+i
    ww.cell(r,1,w); data_cell(ww.cell(r,1), num=False, band=(i%2==1))
    for j,c in enumerate(COMPS):
        ww.cell(r,2+j, ward_comp(w,c)); data_cell(ww.cell(r,2+j), band=(i%2==1))
    ww.cell(r,6, f"=SUM(B{r}:E{r})"); data_cell(ww.cell(r,6), total=True)
    ww.cell(r,7, f"=IF(SUM($F$3:$F$13)=0,0,F{r}/SUM($F$3:$F$13)*100)"); data_cell(ww.cell(r,7), band=(i%2==1)); ww.cell(r,7).number_format=PCTFMT
tr=3+len(WARDS)
ww.cell(tr,1,"Total"); data_cell(ww.cell(tr,1), num=False, total=True)
for j in range(len(COMPS)+1):
    col=get_column_letter(2+j); ww.cell(tr,2+j, f"=SUM({col}3:{col}{tr-1})"); data_cell(ww.cell(tr,2+j), total=True)
ww.cell(tr,7,"=100"); data_cell(ww.cell(tr,7), total=True); ww.cell(tr,7).number_format=PCTFMT
ww.column_dimensions["A"].width=22
for j in range(1,7): ww.column_dimensions[get_column_letter(1+j)].width=11
ww.freeze_panes="B3"
wbar=BarChart(); wbar.type="bar"; wbar.grouping="stacked"; wbar.overlap=100
wbar.title="Issued by ward"; wbar.height=11; wbar.width=18
wbar.add_data(Reference(ww,min_col=2,max_col=5,min_row=2,max_row=2+len(WARDS)), titles_from_data=True)
wbar.set_categories(Reference(ww,min_col=1,min_row=3,max_row=2+len(WARDS)))
for i,c in enumerate(COMPS): wbar.series[i].graphicalProperties.solidFill=COMPFILL[c]
finish_chart(wbar, xt="Ward / Floor", yt="Units issued")
ww.add_chart(wbar,"I2")

# ================= By Month =================
wm = wb.create_sheet("By Month")
title(wm, "Units issued by month & component", 8)
hdr=["Month"]+COMPS+["Received XM","Returned","Total issued"]
for j,h in enumerate(hdr):
    c=wm.cell(2,1+j,h); style_header_cell(c, fill=(COMPFILL[h] if h in COMPFILL else RED))
for i in range(MONTHS_RESERVED):
    r=3+i
    k = KEYS[i] if i < len(KEYS) else None
    wm.cell(r,1, klabel(k) if k else None); data_cell(wm.cell(r,1), num=False, band=(i%2==1))
    for j,c in enumerate(COMPS):
        wm.cell(r,2+j, (issue_month(k,c) if k else None)); data_cell(wm.cell(r,2+j), band=(i%2==1))
    wm.cell(r,6, (section_month(k,"received") if k else None)); data_cell(wm.cell(r,6), band=(i%2==1))
    wm.cell(r,7, (section_month(k,"returned_ward")+section_month(k,"returned_ash") if k else None)); data_cell(wm.cell(r,7), band=(i%2==1))
    wm.cell(r,8, f"=SUM(B{r}:E{r})" if k else None); data_cell(wm.cell(r,8), total=True)
wm.column_dimensions["A"].width=12
for j in range(1,8): wm.column_dimensions[get_column_letter(1+j)].width=12
wm.freeze_panes="B3"
line=LineChart(); line.title="Monthly issuance by component"; line.height=9; line.width=20; line.smooth=False
line.add_data(Reference(wm,min_col=2,max_col=5,min_row=2,max_row=2+len(KEYS)), titles_from_data=True)
line.set_categories(Reference(wm,min_col=1,min_row=3,max_row=2+len(KEYS)))
for i,c in enumerate(COMPS):
    line.series[i].graphicalProperties.line.solidFill=COMPFILL[c]
    line.series[i].graphicalProperties.line.width=20000
    line.series[i].smooth=False
finish_chart(line, xt="Month", yt="Units issued")
wm.add_chart(line,"J2")

out_xlsx = os.path.join(ROOT, "tools/template.xlsx")
wb.save(out_xlsx)
raw = open(out_xlsx,"rb").read()
b64 = base64.b64encode(raw).decode()
open(os.path.join(ROOT,"template-embed.js"),"w").write(
    "// Auto-generated styled Excel template (openpyxl). The app injects live\n"
    "// values into this workbook's cells via JSZip, preserving styles + charts.\n"
    "window.BB_TEMPLATE_B64 = \""+b64+"\";\n"
    "window.BB_TEMPLATE_META = "+json.dumps({
        "yearsReserved":YEARS_RESERVED,"monthsReserved":MONTHS_RESERVED,
        "wards":len(WARDS),"comps":len(COMPS)})+";\n")
print(f"template.xlsx {len(raw)} bytes → template-embed.js base64 {len(b64)} chars")
print("years:",YEARS,"months:",len(KEYS),"grand issued:",section_month_all)
