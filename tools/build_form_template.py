#!/usr/bin/env python3
"""Build ONE professionally styled blank monthly-form worksheet (Form-LAB-ARH-GEN-016
layout: sections x wards/components/tests, days 1..31, Total formulas, a per-day
staff-signature row) and base64-embed it as template-form-embed.js, together with a
BB_FORM_LAYOUT map so the web app can inject each month's values into a clone of the
sheet via JSZip (preserving styling + formulas).
"""
import json, os, base64
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_js = open(os.path.join(ROOT, "data-seed.js")).read()
seed = json.loads(_js[_js.index("{"):_js.rindex("}")+1])
COMPS = seed["components"]                 # PRBC, FFP, CRYO, Platelets
WARDS = seed["wards"]                       # 11 wards
RET   = ["PRBC","Platelets","FFP","CRYO"]
LAB   = ["NO,SPECIMENS REC'D (IP)","NO,SPECIMENS REC'D (OP)","ABO& RhD (IP)","ABO& RhD (OP)",
    "Rh Weak D (OP)","Rh Weak D (IP)","DIRECT COOMBS (OP)","DIRECT COOMBS (IP)","Ab SCREENING RT/37 (IP)",
    "Ab SCREENING RT/37 (OP)","PANEL (OP)","PANEL (IP)","TITRATION (IP)","TITRATION (OP)","Ag TYPING RT (OP)",
    "Ag TYPING RT (IP)","Ag TYPING w/ AHG (IP)","Ag TYPING w/ AHG (OP)","X-MATCHING (IS)","X-MATCHING (AHG)",
    "Trnsfusion reaction","ELUTION","Adsorption"]

NDAYS = 31
DAY_C0 = 2                                  # first day column = B
TOTAL_C = DAY_C0 + NDAYS                     # AG
LAST_C = TOTAL_C

RED="C0392B"; SLATE="34424F"; GREY="6B7280"; LINE="D9D6CE"; BAND="F5F3EF"; TOTALF="EFEBE3"
COMPFILL={"PRBC":"E34948","FFP":"EDA100","CRYO":"2A78D6","Platelets":"1BAF7A"}
thin=Side(style="thin",color=LINE); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)

wb=openpyxl.Workbook(); ws=wb.active; ws.title="FORM"

def cell(r,c,v=None):
    x=ws.cell(r,c,v); x.border=BORDER; return x
def band_header(r, text, fill=SLATE):
    # label cell + day-number cells 1..31 + Total, as a colored header band
    a=cell(r,1,text); a.font=Font(bold=True,color="FFFFFF",size=10); a.fill=PatternFill("solid",fgColor=fill)
    a.alignment=Alignment(horizontal="left",vertical="center",indent=1)
    for d in range(1,NDAYS+1):
        c=cell(r,DAY_C0+d-1,d); c.font=Font(bold=True,color="FFFFFF",size=9)
        c.fill=PatternFill("solid",fgColor=fill); c.alignment=Alignment(horizontal="center",vertical="center")
    t=cell(r,TOTAL_C,"Total"); t.font=Font(bold=True,color="FFFFFF",size=9)
    t.fill=PatternFill("solid",fgColor=fill); t.alignment=Alignment(horizontal="center",vertical="center")
def data_row(r, label, band=False, total=True):
    a=cell(r,1,label); a.alignment=Alignment(horizontal="left",vertical="center",indent=1)
    a.font=Font(size=10, bold=False)
    if band: a.fill=PatternFill("solid",fgColor=BAND)
    for d in range(1,NDAYS+1):
        c=cell(r,DAY_C0+d-1); c.number_format='#,##0'; c.alignment=Alignment(horizontal="center")
        if band: c.fill=PatternFill("solid",fgColor=BAND)
    if total:
        tc=cell(r,TOTAL_C, f"=SUM({get_column_letter(DAY_C0)}{r}:{get_column_letter(TOTAL_C-1)}{r})")
        tc.font=Font(bold=True); tc.fill=PatternFill("solid",fgColor=TOTALF); tc.number_format='#,##0'
    else:
        cell(r,TOTAL_C)

# ---- title ----
ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=LAST_C)
tc=ws.cell(1,1,"Blood Bank Daily Statistics"); tc.font=Font(bold=True,color="FFFFFF",size=13)
tc.fill=PatternFill("solid",fgColor=RED); tc.alignment=Alignment(horizontal="left",vertical="center",indent=1)
ws.row_dimensions[1].height=24
layout={"title":{"row":1,"ref":"A1"},
        "dayCols":{str(d):get_column_letter(DAY_C0+d-1) for d in range(1,NDAYS+1)},
        "totalCol":get_column_letter(TOTAL_C),"sections":[]}

r=3
def add_section(kind, header, rows, comp=None, fill=SLATE, total=True):
    global r
    band_header(r, header, fill=fill); r+=1
    recs=[]
    for i,lab in enumerate(rows):
        data_row(r, lab, band=(i%2==1), total=total); recs.append({"label":lab,"row":r}); r+=1
    sec={"kind":kind,"rows":recs}
    if comp: sec["comp"]=comp
    layout["sections"].append(sec)
    r+=1   # spacer

for c in COMPS:
    add_section("issue", "ISSUING "+c, WARDS, comp=c, fill=COMPFILL[c])
add_section("received", "RECEIVED CROSS MATCHED", WARDS, fill=SLATE)
add_section("returned_ward", "RETURNED FROM WARD", RET, fill=SLATE)
add_section("returned_ash", "RETURNED ARH → ASH", RET, fill=SLATE)
add_section("inventory", "DAILY INVENTORY FROM ASH", RET, fill=SLATE)
add_section("labtests", "TRANSFUSION LAB — NAME OF TEST", LAB, fill=SLATE)

# ---- staff signature row (text per day, no total) ----
band_header(r, "STAFF SIGNATURE", fill=GREY); sig_r=r+1
a=cell(sig_r,1,"Signature"); a.font=Font(bold=True,size=10); a.alignment=Alignment(indent=1,vertical="center")
for d in range(1,NDAYS+1):
    c=cell(sig_r,DAY_C0+d-1); c.alignment=Alignment(horizontal="left",vertical="center"); c.font=Font(size=9)
cell(sig_r,TOTAL_C)
layout["sigRow"]=sig_r

# ---- widths / freeze ----
ws.column_dimensions["A"].width=26
for d in range(1,NDAYS+1): ws.column_dimensions[get_column_letter(DAY_C0+d-1)].width=4.2
ws.column_dimensions[get_column_letter(TOTAL_C)].width=7.5
ws.freeze_panes="B3"

out_xlsx=os.path.join(ROOT,"tools/template_form.xlsx"); wb.save(out_xlsx)
raw=open(out_xlsx,"rb").read(); b64=base64.b64encode(raw).decode()
open(os.path.join(ROOT,"template-form-embed.js"),"w").write(
    "// Auto-generated styled blank monthly-form template (openpyxl). The app clones\n"
    "// this sheet per month and injects values via JSZip, preserving styles+formulas.\n"
    "window.BB_FORM_TEMPLATE_B64 = \""+b64+"\";\n"
    "window.BB_FORM_LAYOUT = "+json.dumps(layout,separators=(',',':'))+";\n")
print(f"template_form.xlsx {len(raw)} bytes; rows used up to {sig_r}; sections {len(layout['sections'])}; sigRow {sig_r}")
print("base64", len(b64), "chars")
